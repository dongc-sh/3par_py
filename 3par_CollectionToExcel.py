import urllib3
from hpe3parclient import client
import pandas as pd
from openpyxl import load_workbook


def Connect3par(hpe3parip, username, password):
    cl = client.HPE3ParClient('https://%s:8080/api/v1' % hpe3parip)
    cl.setSSHOptions(ip=hpe3parip, login=username, password=password)
    cl.login(username, password)
    return cl

def GetSysteminfo(cl, *systeminfo):
    systeminfo = cl.getStorageSystemInfo()
    systemVersion.append(systeminfo['systemVersion'])
    storagename.append(systeminfo['name'])
    storagemodel.append(systeminfo['model'])
    Serialnumber.append(systeminfo['serialNumber'])
    MgmtIP.append(hpe3parip)
    return systeminfo

def GetCapactiy(cl, *param):
    capacity = cl.getOverallSystemCapacity()
    SSDtotal.append(int(capacity['SSDCapacity']['totalMiB'])/1024)
    SSDFree.append(int(capacity['SSDCapacity']['freeMiB'])/1024)
    FCtotal.append(int(capacity['FCCapacity']['totalMiB']) / 1024)
    FCFree.append(int(capacity['FCCapacity']['freeMiB']) / 1024)
    NLtotal.append(int(capacity['NLCapacity']['totalMiB']) / 1024)
    NLFree.append(int(capacity['NLCapacity']['freeMiB']) / 1024)
    return param


def ExcelMod(workbook_path):
    workbook = load_workbook(path)
    worksheet = workbook[workbook.sheetnames[0]]
    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        worksheet.column_dimensions[i].width = 13                                        #修改excel表格列宽
    workbook.save(workbook_path)

if __name__ == '__main__':
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    dev_filepath = r"C:\\users\c64146\PycharmProjects\HP3par\device_info.txt"
    dev_file = open(dev_filepath, "r")
    systemVersion,storagename,storagemodel,Serialnumber,MgmtIP = [[] for x in range(5)]
    SSDtotal,SSDFree,FCtotal,FCFree,NLtotal,NLFree = [[] for x in range(6)]                      #定义空数组，用于存放收集的数据


    while 1:
        dev_info = dev_file.readline()
        if not dev_info:
            break
        else:
            devs = dev_info.split(',')
            hpe3parip = devs[0]
            username = devs[1]
            password = devs[2].strip()
            password = password.strip('\n')

            conn = Connect3par(hpe3parip, username, password)
            GetSysteminfo(conn,systemVersion,storagename,storagemodel,Serialnumber,MgmtIP)
            GetCapactiy(conn,SSDtotal,SSDFree,FCtotal,FCFree,NLtotal,NLFree)

            conn.logout()

    df_systeminfo = pd.DataFrame({'HostName':storagename,'Model':storagemodel,'Version':systemVersion,'SN':Serialnumber,'Mgmt IP':MgmtIP})
    df_capacity = pd.DataFrame({'SSD total(GB)':SSDtotal,'SSD free(GB)':SSDFree,'FC total(GB)':FCtotal,'FC free(GB)':FCFree,'NL total(GB)':NLtotal,'NL free(GB)':NLFree})
    output = pd.concat([df_systeminfo, df_capacity], axis=1)                 #合并两个dataframe
    output.to_excel('output.xlsx', index=False, sheet_name="storageinfo")    #将合并后的df导出至excel表格
    path = 'output.xlsx'
    ExcelMod(path)                                                      #修改excel表格列宽

    dev_file.close()

